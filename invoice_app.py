import json
import logging
import os
import re
import tempfile
from datetime import datetime
from io import BytesIO
from typing import Any
from urllib.parse import urlparse
from uuid import uuid4

import fitz
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
try:
    import msal
except ModuleNotFoundError:
    msal = None
import pandas as pd
import requests
import streamlit as st
from dotenv import load_dotenv
from google.cloud import vision
from google.oauth2 import service_account
from openai import APIStatusError, OpenAI, RateLimitError

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
logger = logging.getLogger(__name__)

load_dotenv()

COMPANY_OPTIONS = [
    "1001298527 ONTARIO INC",
    "10342548 CANADA INC",
    "10696480 CANADA LTD",
    "12433087 CANADA INC-MASTER",
    "13037622 CANADA INC",
    "9359-6633 QUEBEC INC",
    "9390-9216 QUEBEC INC",
    "D-TECH CONSTRUCTION",
    "TAYANTI-CANADA",
]
BANK_OPTIONS = ["Scotiabank", "Desjardins", "National Bank"]
ALLOWED_CATEGORIES = {"gas", "parking", "meals", "supplier", "other"}
MICROSOFT_SCOPES = ["User.Read", "Files.Read.All"]
MSAL_SESSION_CACHE_KEY = "msal_token_cache_serialized"
DEFAULT_RECEIPTS_DATABASE_DIR = (
    "General/Sales receipts database"
)
RECEIPTS_DATABASE_CSV = "sales_receipts_database.csv"
PROVINCE_CODES = {
    "AB",
    "BC",
    "MB",
    "NB",
    "NL",
    "NS",
    "NT",
    "NU",
    "ON",
    "PE",
    "QC",
    "SK",
    "YT",
}


def _get_config_value(key: str, default: str = "") -> str:
    try:
        if key in st.secrets:
            return str(st.secrets[key]).strip()
    except Exception:
        pass
    return os.getenv(key, default).strip()


def _to_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _safe_json(data: Any) -> str:
    try:
        return json.dumps(data, ensure_ascii=False)
    except Exception:
        return "{}"


def _sanitize_filename_component(value: str, fallback: str) -> str:
    text = str(value or "").strip()
    if not text:
        return fallback
    text = re.sub(r"[^\w\-]+", "_", text, flags=re.UNICODE)
    text = re.sub(r"_+", "_", text).strip("_")
    return text or fallback


def _normalize_card_last4(value: Any) -> str:
    text = str(value or "").strip().upper()
    if text.startswith("N"):
        text = text[1:]
    if not re.fullmatch(r"\d{4}", text):
        return ""
    return f"N{text}"


def _display_card_last4(value: Any) -> str:
    normalized = _normalize_card_last4(value)
    return normalized[1:] if normalized else ""


def _normalize_card_last4_series(series: pd.Series) -> pd.Series:
    return series.map(_normalize_card_last4).astype("string")


def _microsoft_auth_available() -> tuple[bool, str]:
    if msal is None:
        return False, "Missing dependency: msal"

    tenant_id = _get_config_value("TENANT_ID")
    client_id = _get_config_value("CLIENT_ID")
    client_secret = _get_config_value("CLIENT_SECRET")
    if not tenant_id or not client_id or not client_secret:
        return False, "Missing TENANT_ID, CLIENT_ID or CLIENT_SECRET in secrets/environment"

    return True, ""


def _load_msal_cache():
    cache = msal.SerializableTokenCache()
    serialized = st.session_state.get(MSAL_SESSION_CACHE_KEY)
    if serialized:
        cache.deserialize(str(serialized))
    return cache


def _save_msal_cache(cache) -> None:
    if cache.has_state_changed:
        st.session_state[MSAL_SESSION_CACHE_KEY] = cache.serialize()


def _resolve_redirect_uri() -> str:
    env_redirect = _get_config_value("REDIRECT_URI")
    host = ""
    try:
        host = str(st.context.headers.get("host", "")).strip()
    except Exception:
        host = ""

    if host:
        is_local = host.startswith("localhost") or host.startswith("127.0.0.1")
        inferred = f"{'http' if is_local else 'https'}://{host}/"
        if env_redirect:
            try:
                env_host = urlparse(env_redirect).netloc.lower()
            except Exception:
                env_host = ""
            if env_host and env_host != host.lower():
                return inferred
            return env_redirect
        return inferred

    if env_redirect:
        return env_redirect
    raise RuntimeError("Missing REDIRECT_URI in environment.")


def _msal_confidential_app(cache):
    tenant_id = _get_config_value("TENANT_ID")
    client_id = _get_config_value("CLIENT_ID")
    client_secret = _get_config_value("CLIENT_SECRET")
    if not tenant_id or not client_id or not client_secret:
        raise RuntimeError("Missing TENANT_ID / CLIENT_ID / CLIENT_SECRET in secrets/environment.")

    authority = f"https://login.microsoftonline.com/{tenant_id}"
    return msal.ConfidentialClientApplication(
        client_id,
        authority=authority,
        client_credential=client_secret,
        token_cache=cache,
    )


def get_microsoft_token_silent() -> str | None:
    available, _ = _microsoft_auth_available()
    if not available:
        return None

    cache = _load_msal_cache()
    app = _msal_confidential_app(cache)
    accounts = app.get_accounts()
    if not accounts:
        return None

    result = app.acquire_token_silent(MICROSOFT_SCOPES, account=accounts[0])
    if result and "access_token" in result:
        _save_msal_cache(cache)
        return result["access_token"]
    return None


def get_microsoft_login_url(state: str) -> str:
    available, reason = _microsoft_auth_available()
    if not available:
        raise RuntimeError(reason)

    cache = _load_msal_cache()
    app = _msal_confidential_app(cache)
    redirect_uri = _resolve_redirect_uri()
    return app.get_authorization_request_url(
        MICROSOFT_SCOPES,
        redirect_uri=redirect_uri,
        state=state,
        prompt="select_account",
    )


def finish_microsoft_redirect_flow(auth_code: str) -> str:
    available, reason = _microsoft_auth_available()
    if not available:
        raise RuntimeError(reason)

    cache = _load_msal_cache()
    app = _msal_confidential_app(cache)
    redirect_uri = _resolve_redirect_uri()
    result = app.acquire_token_by_authorization_code(
        code=auth_code,
        scopes=MICROSOFT_SCOPES,
        redirect_uri=redirect_uri,
    )
    if "access_token" not in result:
        raise RuntimeError(str(result))

    _save_msal_cache(cache)
    return result["access_token"]


def clear_microsoft_session() -> None:
    st.session_state.pop("graph_token", None)
    st.session_state.pop("oauth_state", None)
    st.session_state.pop(MSAL_SESSION_CACHE_KEY, None)


def build_suggested_file_name(payment_date: str, bank: str, card_type: str, merchant_name: str, total_amount: float) -> str:
    date_part_match = re.search(r"\d{4}-\d{2}-\d{2}", str(payment_date or ""))
    date_part = date_part_match.group(0) if date_part_match else str(payment_date or "")
    amount_part = f"{_to_float(total_amount, 0.0):.2f}"
    parts = [
        _sanitize_filename_component(date_part, "no_date"),
        _sanitize_filename_component(bank, "no_bank"),
        _sanitize_filename_component(card_type, "no_card"),
        _sanitize_filename_component(merchant_name, "no_merchant"),
        _sanitize_filename_component(amount_part, "0_00"),
    ]
    return "__".join(parts)


def _database_dir() -> str:
    return _get_config_value("RECEIPTS_DATABASE_DIR", DEFAULT_RECEIPTS_DATABASE_DIR)


def _database_csv_path() -> str:
    return f"{_database_dir().rstrip('/')}/{RECEIPTS_DATABASE_CSV}"


def _join_sp_path(*parts: str) -> str:
    cleaned = [str(part or "").strip().strip("/") for part in parts if str(part or "").strip()]
    return "/".join(cleaned)


def extract_pdf_page_bytes(pdf_bytes: bytes, page_index: int) -> bytes:
    with fitz.open(stream=pdf_bytes, filetype="pdf") as src:
        if page_index < 0 or page_index >= src.page_count:
            raise IndexError(f"Page index out of range: {page_index}")
        single = fitz.open()
        single.insert_pdf(src, from_page=page_index, to_page=page_index)
        try:
            return single.tobytes(garbage=4, deflate=True)
        finally:
            single.close()


def build_database_row(
    *,
    page_number: int,
    company: str,
    bank: str,
    card_type: str,
    card_last4: str,
    gpt_json: dict[str, Any],
    notes: str,
    pdf_file_name: str,
    pdf_file_path: str,
) -> dict[str, Any]:
    return {
        "processed_at": datetime.now().isoformat(timespec="seconds"),
        "source_page_number": int(page_number),
        "company": company,
        "bank": bank,
        "card_type": card_type,
        "card_last4": card_last4,
        "gpt_payment_date": str(gpt_json.get("payment_date") or "").strip(),
        "gpt_total_amount": _to_float(gpt_json.get("total_amount"), 0.0),
        "gpt_taxes_total": _to_float(gpt_json.get("taxes_total"), 0.0),
        "gpt_category": str(gpt_json.get("category") or "other").strip().lower(),
        "gpt_merchant_name": str(gpt_json.get("merchant_name") or "").strip(),
        "gpt_city": str(gpt_json.get("city") or "").strip(),
        "gpt_province": str(gpt_json.get("province") or "").strip(),
        "gpt_confidence": _to_float(gpt_json.get("confidence"), 0.0),
        "notes": notes,
        "file_name": pdf_file_name,
        "file_path": pdf_file_path,
    }


def graph_get(url: str, token: str) -> dict[str, Any]:
    response = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=60)
    if response.status_code >= 400:
        raise RuntimeError(response.text)
    return response.json()


def graph_download(url: str, token: str) -> bytes:
    response = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=120)
    if response.status_code >= 400:
        raise RuntimeError(response.text)
    return response.content


def graph_put_bytes(url: str, token: str, content: bytes) -> dict[str, Any]:
    response = requests.put(url, headers={"Authorization": f"Bearer {token}"}, data=content, timeout=120)
    if response.status_code >= 400:
        raise RuntimeError(response.text)
    return response.json() if response.content else {}


def resolve_drive_id(token: str) -> str:
    sp_hostname = _get_config_value("SP_HOSTNAME")
    sp_site_path = _get_config_value("SP_SITE_PATH")
    sp_drive_name = _get_config_value("SP_DRIVE_NAME", "Documents")
    if not sp_hostname or not sp_site_path:
        raise RuntimeError("Missing SP_HOSTNAME / SP_SITE_PATH in secrets/environment.")

    site = graph_get(f"https://graph.microsoft.com/v1.0/sites/{sp_hostname}:{sp_site_path}", token)
    drives = graph_get(f"https://graph.microsoft.com/v1.0/sites/{site['id']}/drives", token).get("value", [])
    drive = next((d for d in drives if d.get("name") == sp_drive_name), None) or (drives[0] if drives else None)
    if not drive:
        raise RuntimeError("Could not resolve SharePoint drive.")
    return str(drive["id"])


def list_children_by_path(drive_id: str, sp_path: str, token: str) -> list[dict[str, Any]]:
    encoded_path = requests.utils.quote(sp_path.strip("/"), safe="/")
    url = (
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{encoded_path}:/children"
        f"?$top=200&$select=id,name,folder,file,webUrl"
    )
    items: list[dict[str, Any]] = []
    while url:
        data = graph_get(url, token)
        items.extend(data.get("value", []))
        url = data.get("@odata.nextLink")
    return items


def download_sharepoint_file_bytes(sp_relative_path: str, token: str, drive_id: str | None = None) -> bytes:
    did = drive_id or resolve_drive_id(token)
    encoded_path = requests.utils.quote(sp_relative_path.strip("/"), safe="/")
    url = f"https://graph.microsoft.com/v1.0/drives/{did}/root:/{encoded_path}:/content"
    return graph_download(url, token)


def upload_sharepoint_file_bytes(
    sp_relative_path: str,
    content: bytes,
    token: str,
    drive_id: str | None = None,
) -> dict[str, Any]:
    did = drive_id or resolve_drive_id(token)
    encoded_path = requests.utils.quote(sp_relative_path.strip("/"), safe="/")
    url = f"https://graph.microsoft.com/v1.0/drives/{did}/root:/{encoded_path}:/content"
    return graph_put_bytes(url, token, content)


def load_database_df(token: str, drive_id: str) -> pd.DataFrame:
    csv_path = _database_csv_path()
    try:
        content = download_sharepoint_file_bytes(csv_path, token, drive_id=drive_id)
    except Exception:
        return pd.DataFrame()
    try:
        df = pd.read_csv(BytesIO(content), dtype={"card_last4": "string"})
        if "card_last4" in df.columns:
            df["card_last4"] = _normalize_card_last4_series(df["card_last4"])
        return df
    except Exception as exc:
        logger.exception("Could not read database CSV")
        st.error(f"Could not read database CSV: {exc}")
        return pd.DataFrame()


def append_rows_to_database(rows_df: pd.DataFrame, token: str, drive_id: str) -> str:
    csv_path = _database_csv_path()
    existing_df = load_database_df(token, drive_id)
    combined_df = pd.concat([existing_df, rows_df], ignore_index=True) if not existing_df.empty else rows_df.copy()
    if "card_last4" in combined_df.columns:
        combined_df["card_last4"] = _normalize_card_last4_series(combined_df["card_last4"])
    buffer = BytesIO()
    combined_df.to_csv(buffer, index=False, encoding="utf-8-sig")
    upload_sharepoint_file_bytes(csv_path, buffer.getvalue(), token, drive_id=drive_id)
    return csv_path


def existing_remote_file_names(token: str, drive_id: str) -> set[str]:
    try:
        items = list_children_by_path(drive_id, _database_dir(), token)
    except Exception:
        return set()
    return {str(item.get("name") or "").strip() for item in items if str(item.get("name") or "").strip()}


def make_unique_remote_pdf_name(base_name: str, existing_names: set[str]) -> tuple[str, str]:
    safe_base = _sanitize_filename_component(base_name, "invoice")
    candidate_name = f"{safe_base}.pdf"
    counter = 2
    while candidate_name in existing_names:
        candidate_name = f"{safe_base}__{counter}.pdf"
        counter += 1
    existing_names.add(candidate_name)
    return candidate_name, _join_sp_path(_database_dir(), candidate_name)


def merge_pdf_files(pdf_blobs: list[bytes]) -> bytes:
    merged = fitz.open()
    try:
        for pdf_blob in pdf_blobs:
            if not pdf_blob:
                continue
            with fitz.open(stream=pdf_blob, filetype="pdf") as src:
                merged.insert_pdf(src)
        return merged.tobytes(garbage=4, deflate=True)
    finally:
        merged.close()


def render_pdf_pages_to_png_paths(pdf_bytes: bytes, output_dir: str, dpi: int = 200) -> list[dict[str, Any]]:
    if not pdf_bytes:
        return []
    os.makedirs(output_dir, exist_ok=True)
    zoom = dpi / 72.0
    matrix = fitz.Matrix(zoom, zoom)
    pages: list[dict[str, Any]] = []
    with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
        for idx, page in enumerate(doc, start=1):
            pix = page.get_pixmap(matrix=matrix, alpha=False)
            path = os.path.join(output_dir, f"page_{idx}.png")
            pix.save(path)
            pages.append({"page_number": idx, "image_path": path})
    return pages


def get_google_vision_client() -> vision.ImageAnnotatorClient:
    if "google_service_account" not in st.secrets:
        raise RuntimeError("Missing google_service_account in Streamlit secrets.")

    credentials_info = dict(st.secrets["google_service_account"])
    credentials = service_account.Credentials.from_service_account_info(credentials_info)
    return vision.ImageAnnotatorClient(credentials=credentials)


def extract_text_google_vision(image_bytes: bytes) -> dict[str, Any]:
    client = get_google_vision_client()
    image = vision.Image(content=image_bytes)
    response = client.document_text_detection(image=image)
    if response.error and response.error.message:
        raise RuntimeError(f"Google Vision error: {response.error.message}")
    full_text = ""
    if response.full_text_annotation and response.full_text_annotation.text:
        full_text = response.full_text_annotation.text
    return {"full_text": full_text}


def _extract_amount_from_line(line: str) -> float:
    # Accept 1,234.56 or 1234.56
    candidates = re.findall(r"(?<!\d)(\d{1,3}(?:,\d{3})*(?:\.\d{2})|\d+\.\d{2})(?!\d)", line)
    if not candidates:
        return 0.0
    return max(_to_float(c.replace(",", ""), 0.0) for c in candidates)


def parse_receipt_from_text(full_text: str) -> dict[str, Any]:
    lines = [ln.strip() for ln in (full_text or "").splitlines() if ln.strip()]
    merchant_name = lines[0] if lines else ""

    date_iso = ""
    date_match = re.search(
        r"\b(20\d{2}[-/](0[1-9]|1[0-2])[-/](0[1-9]|[12]\d|3[01]))(?:[ T]([01]\d|2[0-3]):[0-5]\d(?::[0-5]\d)?)?\b",
        full_text,
    )
    if date_match:
        base = date_match.group(1).replace("/", "-")
        time_part = date_match.group(4) or "00:00:00"
        if len(time_part) == 5:
            time_part = f"{time_part}:00"
        date_iso = f"{base}T{time_part}"

    total_amount = 0.0
    taxes_total = 0.0
    for line in lines:
        upper = line.upper()
        amount = _extract_amount_from_line(line)
        if amount <= 0:
            continue
        if "TOTAL" in upper and "SUBTOTAL" not in upper:
            total_amount = max(total_amount, amount)
        if any(tax_kw in upper for tax_kw in ("TAX", "GST", "HST", "TVQ", "TPS", "VAT")):
            taxes_total += amount

    if total_amount == 0.0:
        # fallback: biggest amount in receipt
        for line in lines:
            total_amount = max(total_amount, _extract_amount_from_line(line))

    city = ""
    province = ""
    for line in lines:
        prov_match = re.search(r"\b([A-Z]{2})\b", line.upper())
        if prov_match and prov_match.group(1) in PROVINCE_CODES:
            province = prov_match.group(1)
            city_candidate = line[: prov_match.start()].strip(" ,.-")
            if city_candidate:
                city = city_candidate.split(",")[-1].strip()
            break

    compact = {
        "merchant": merchant_name,
        "date": date_iso,
        "total": total_amount,
        "taxes_total": taxes_total,
        "city": city,
        "province": province,
        "address": "",
        "items": [],
    }
    return compact


def _truncate_notes(text: str) -> str:
    words = (text or "").split()
    return text if len(words) <= 20 else " ".join(words[:20])


def fallback_gpt(reason: str) -> dict[str, Any]:
    return {
        "payment_date": "",
        "total_amount": 0.0,
        "taxes_total": 0.0,
        "category": "other",
        "confidence": 0.0,
        "merchant_name": "",
        "city": "",
        "province": "",
        "notes": _truncate_notes(f"Fallback: {reason}"),
    }


def _trim_text(text: str, max_len: int = 12000) -> str:
    value = str(text or "")
    if len(value) <= max_len:
        return value
    return value[:max_len]


def _normalize_gpt_output(parsed: dict[str, Any]) -> dict[str, Any]:
    category = str(parsed.get("category", "other")).strip().lower()
    if category not in ALLOWED_CATEGORIES:
        category = "other"
    return {
        "payment_date": str(parsed.get("payment_date", "") or "").strip(),
        "total_amount": _to_float(parsed.get("total_amount"), 0.0),
        "taxes_total": _to_float(parsed.get("taxes_total"), 0.0),
        "category": category,
        "confidence": max(0.0, min(1.0, _to_float(parsed.get("confidence"), 0.0))),
        "merchant_name": str(parsed.get("merchant_name", "") or "").strip(),
        "city": str(parsed.get("city", "") or "").strip(),
        "province": str(parsed.get("province", "") or "").strip(),
        "notes": _truncate_notes(str(parsed.get("notes", "") or "").strip()),
    }


def _missing_text_fields(gpt_json: dict[str, Any]) -> list[str]:
    required = ["payment_date", "merchant_name", "city", "province", "category"]
    missing: list[str] = []
    for key in required:
        if str(gpt_json.get(key) or "").strip() == "":
            missing.append(key)
    return missing


def _single_gpt_call(client: OpenAI, system_prompt: str, user_prompt: str) -> dict[str, Any]:
    response = client.chat.completions.create(
        model="gpt-4o-mini",
        temperature=0,
        response_format={"type": "json_object"},
        messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}],
    )
    content = (response.choices[0].message.content or "").strip()
    parsed = json.loads(content)
    if not isinstance(parsed, dict):
        raise ValueError("GPT did not return a JSON object")
    return _normalize_gpt_output(parsed)


def classify_with_gpt(
    vision_json: dict[str, Any],
    compact_receipt: dict[str, Any],
    use_location_enrichment: bool = True,
) -> dict[str, Any]:
    api_key = _get_config_value("OPENAI_API_KEY")
    if not api_key:
        return fallback_gpt("OPENAI_API_KEY is missing in secrets/environment")

    client = OpenAI(api_key=api_key)
    enrich_rule = (
        "If city/province are missing, infer them from merchant/address text."
        if use_location_enrichment
        else "Do not infer city/province. Keep them blank if unavailable."
    )
    system_prompt = (
        "You are extracting fields from an invoice OCR result. "
        "Return JSON only with keys: payment_date, total_amount, taxes_total, category, confidence, "
        "merchant_name, city, province, notes. Category must be one of gas, parking, meals, supplier, other. "
        "payment_date must be ISO datetime when possible. total_amount and taxes_total must be numbers."
    )
    vision_payload = dict(vision_json or {})
    vision_payload["full_text"] = _trim_text(vision_payload.get("full_text", ""))
    required_fields = [
        "payment_date",
        "total_amount",
        "taxes_total",
        "category",
        "merchant_name",
        "city",
        "province",
    ]
    today = datetime.now()
    current_year = today.year
    previous_year = current_year - 1
    user_prompt = (
        "Esto viene de un invoice. Quiero que llenes estos campos: "
        f"{', '.join(required_fields)}. "
        f"{enrich_rule} Prefer OCR evidence. If uncertain category use other and lower confidence. Notes <=20 words. "
        f"For payment_date, the invoice year is very likely {current_year}; depending on how close the receipt is to the start of the year, it may be {previous_year} instead.\n"
        f"Google Vision OCR JSON:\n{json.dumps(vision_payload, ensure_ascii=False)}\n\n"
        f"Parsed helper fields:\n{json.dumps(compact_receipt, ensure_ascii=False)}"
    )
    try:
        merged = _single_gpt_call(client=client, system_prompt=system_prompt, user_prompt=user_prompt)

        # Up to 3 total attempts to fill missing text fields.
        for attempt in range(2, 4):
            missing = _missing_text_fields(merged)
            if not missing:
                break
            retry_prompt = (
                f"{user_prompt}\n\n"
                f"Attempt {attempt}/3. Your previous output left these fields empty: {', '.join(missing)}.\n"
                "Return JSON again and fill every missing field using the OCR evidence. "
                "Do not leave those fields blank."
            )
            retry_result = _single_gpt_call(client=client, system_prompt=system_prompt, user_prompt=retry_prompt)
            for key, val in retry_result.items():
                if str(merged.get(key) or "").strip() == "" and str(val or "").strip() != "":
                    merged[key] = val

        return merged
    except RateLimitError as exc:
        return fallback_gpt(f"OpenAI rate limit/quota: {exc}")
    except APIStatusError as exc:
        return fallback_gpt(f"OpenAI API error: {exc}")
    except Exception as exc:
        return fallback_gpt(f"OpenAI parse/call failed: {exc}")


def create_excel_bytes(summary_df: pd.DataFrame, raw_df: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="summary", index=False)
        raw_df.to_excel(writer, sheet_name="raw", index=False)
    output.seek(0)
    return output.read()


def create_filtered_excel_bytes(filtered_df: pd.DataFrame) -> bytes:
    output = BytesIO()
    export_df = filtered_df.copy()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="receipts", index=False)
        worksheet = writer.sheets["receipts"]

        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style="thin", color="D9E2F3"),
            right=Side(style="thin", color="D9E2F3"),
            top=Side(style="thin", color="D9E2F3"),
            bottom=Side(style="thin", color="D9E2F3"),
        )
        even_fill = PatternFill(fill_type="solid", fgColor="F7FBFF")
        odd_fill = PatternFill(fill_type="solid", fgColor="EAF2F8")

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            fill = even_fill if row_idx % 2 == 0 else odd_fill
            for cell in row:
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for idx, column_name in enumerate(export_df.columns, start=1):
            series = export_df[column_name].astype(str) if column_name in export_df.columns else pd.Series(dtype=str)
            max_length = max([len(str(column_name))] + [len(value) for value in series.head(200).tolist()]) if not series.empty else len(str(column_name))
            worksheet.column_dimensions[get_column_letter(idx)].width = min(max(max_length + 2, 12), 40)

    output.seek(0)
    return output.read()


def _format_column_label(column_name: str) -> str:
    text = str(column_name or "").strip()
    if text.startswith("gpt_"):
        text = text[4:]
    return text.replace("_", " ").title()


def _prettify_dataframe_columns(df: pd.DataFrame) -> pd.DataFrame:
    renamed_df = df.copy()
    renamed_df.columns = [_format_column_label(column) for column in renamed_df.columns]
    return renamed_df


def _apply_date_range_filter(filtered_df: pd.DataFrame, column: str, label: str, key: str) -> pd.DataFrame:
    if column not in filtered_df.columns:
        return filtered_df

    parsed_dates = pd.to_datetime(filtered_df[column], errors="coerce")
    valid_dates = parsed_dates.dropna()
    if valid_dates.empty:
        return filtered_df

    min_date = valid_dates.min().date()
    max_date = valid_dates.max().date()
    selected_range = st.date_input(label, value=(min_date, max_date), min_value=min_date, max_value=max_date, key=key)
    if isinstance(selected_range, tuple) and len(selected_range) == 2:
        start_date, end_date = selected_range
        mask = parsed_dates.dt.date.between(start_date, end_date, inclusive="both")
        return filtered_df[mask.fillna(False)]
    return filtered_df


def _apply_numeric_slider_filter(filtered_df: pd.DataFrame, column: str, label: str, key: str) -> pd.DataFrame:
    if column not in filtered_df.columns:
        return filtered_df

    numeric_series = pd.to_numeric(filtered_df[column], errors="coerce")
    valid_values = numeric_series.dropna()
    if valid_values.empty:
        return filtered_df

    min_value = float(valid_values.min())
    max_value = float(valid_values.max())
    if min_value == max_value:
        return filtered_df

    selected_min, selected_max = st.slider(
        label,
        min_value=min_value,
        max_value=max_value,
        value=(min_value, max_value),
        key=key,
    )
    mask = numeric_series.between(selected_min, selected_max, inclusive="both")
    return filtered_df[mask.fillna(False)]


def render_database_browser(database_df: pd.DataFrame, token: str, drive_id: str, section_key: str = "database") -> None:
    filter_columns = [
        "company",
        "bank",
        "card_type",
        "card_last4",
        "gpt_category",
        "gpt_merchant_name",
        "gpt_city",
        "gpt_province",
    ]
    filtered_df = database_df.copy()
    if "card_last4" in filtered_df.columns:
        filtered_df["card_last4"] = _normalize_card_last4_series(filtered_df["card_last4"])
    with st.sidebar:
        st.header("Filters")
        for column in filter_columns:
            if column not in filtered_df.columns:
                continue
            raw_values = [str(v) for v in filtered_df[column].dropna().astype(str).unique() if str(v).strip()]
            if column == "card_last4":
                option_map = {normalized: normalized for normalized in raw_values}
                options = sorted(option_map)
            else:
                option_map = {}
                options = sorted(raw_values)
            selected = st.multiselect(
                f"Filter by {_format_column_label(column)}",
                options=options,
                key=f"{section_key}_filter_{column}",
            )
            if selected:
                if column == "card_last4":
                    selected_values = [option_map[value] for value in selected if value in option_map]
                    filtered_df = filtered_df[filtered_df[column].astype(str).isin(selected_values)]
                else:
                    filtered_df = filtered_df[filtered_df[column].astype(str).isin(selected)]

        text_query = st.text_input("Search in table", key=f"{section_key}_text_query")
        if text_query.strip():
            query = text_query.strip().lower()
            row_mask = filtered_df.astype(str).apply(
                lambda row: row.str.lower().str.contains(query, na=False).any(),
                axis=1,
            )
            filtered_df = filtered_df[row_mask]

        filtered_df = _apply_date_range_filter(
            filtered_df,
            column="processed_at",
            label="Processed At",
            key=f"{section_key}_filter_processed_at",
        )
        filtered_df = _apply_date_range_filter(
            filtered_df,
            column="gpt_payment_date",
            label="Payment Date",
            key=f"{section_key}_filter_gpt_payment_date",
        )
        filtered_df = _apply_numeric_slider_filter(
            filtered_df,
            column="gpt_total_amount",
            label="Amount",
            key=f"{section_key}_filter_gpt_total_amount",
        )
        filtered_df = _apply_numeric_slider_filter(
            filtered_df,
            column="gpt_taxes_total",
            label="Taxes Total",
            key=f"{section_key}_filter_gpt_taxes_total",
        )
        filtered_df = _apply_numeric_slider_filter(
            filtered_df,
            column="gpt_confidence",
            label="Confidence",
            key=f"{section_key}_filter_gpt_confidence",
        )

    display_df = _prettify_dataframe_columns(filtered_df)
    st.subheader("Filtered Receipts")
    st.caption(f"Rows: {len(display_df)}")
    st.dataframe(
        display_df,
        use_container_width=True,
        hide_index=True,
        column_config={"Card Last4": st.column_config.TextColumn("Card Last4")},
    )

    pdf_paths = filtered_df["file_path"].dropna().astype(str).tolist() if "file_path" in filtered_df.columns else []
    current_signature = "||".join(pdf_paths)
    pdf_bytes_key = f"{section_key}_filtered_pdf_bytes"
    pdf_signature_key = f"{section_key}_filtered_pdf_signature"
    pdf_count_key = f"{section_key}_filtered_pdf_count"
    pdf_skipped_key = f"{section_key}_filtered_pdf_skipped"
    excel_bytes = create_filtered_excel_bytes(display_df)

    st.subheader("Downloads")
    action_col1, action_col2 = st.columns(2)

    with action_col1:
        load_clicked = st.button(
            "Load filtered invoices",
            key=f"{section_key}_load_filtered_invoices",
            disabled=not pdf_paths,
            use_container_width=True,
        )
        if load_clicked:
            pdf_blobs: list[bytes] = []
            missing_count = 0
            for remote_path in pdf_paths:
                try:
                    pdf_blobs.append(download_sharepoint_file_bytes(remote_path, token, drive_id=drive_id))
                except Exception:
                    missing_count += 1

            if pdf_blobs:
                st.session_state[pdf_bytes_key] = merge_pdf_files(pdf_blobs)
                st.session_state[pdf_signature_key] = current_signature
                st.session_state[pdf_count_key] = len(pdf_blobs)
                st.session_state[pdf_skipped_key] = missing_count
                st.success(f"Loaded {len(pdf_blobs)} invoice PDF(s).")
            else:
                st.session_state[pdf_bytes_key] = b""
                st.session_state[pdf_signature_key] = ""
                st.session_state[pdf_count_key] = 0
                st.session_state[pdf_skipped_key] = missing_count
                st.error("No filtered PDFs could be downloaded from SharePoint.")

        if st.session_state.get(pdf_signature_key) == current_signature and st.session_state.get(pdf_bytes_key):
            loaded_count = int(st.session_state.get(pdf_count_key, 0))
            skipped_count = int(st.session_state.get(pdf_skipped_key, 0))
            st.caption(f"Loaded invoices: {loaded_count}")
            if skipped_count:
                st.warning(f"{skipped_count} filtered PDF(s) could not be loaded and were skipped.")
        elif pdf_paths:
            st.caption("Load the current filtered invoices before downloading them.")
        else:
            st.caption("No filtered invoices available to load.")

        st.download_button(
            label="Download loaded invoices",
            data=st.session_state.get(pdf_bytes_key, b""),
            file_name="filtered_invoices_merged.pdf",
            mime="application/pdf",
            key=f"{section_key}_download_loaded_invoices",
            disabled=not (
                st.session_state.get(pdf_signature_key) == current_signature and st.session_state.get(pdf_bytes_key)
            ),
            use_container_width=True,
        )

    with action_col2:
        st.download_button(
            label="Download filtered table in Excel",
            data=excel_bytes,
            file_name="filtered_receipts.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"{section_key}_download_filtered_excel",
            use_container_width=True,
        )


def initialize_session_state() -> None:
    if "gcv_processed" not in st.session_state:
        st.session_state.gcv_processed = False
        st.session_state.gcv_results_saved = False
        st.session_state.gcv_summary_df = pd.DataFrame()
        st.session_state.gcv_summary_preview_df = pd.DataFrame()
        st.session_state.gcv_raw_df = pd.DataFrame()
        st.session_state.gcv_excel_bytes = b""
        st.session_state.gcv_errors = []
        st.session_state.last_database_csv_path = ""
        st.session_state.gcv_pending_uploads = []
    st.session_state.setdefault("gcv_upload_signature", "")
    st.session_state.setdefault("gcv_upload_pdf", None)
    st.session_state.setdefault("gcv_company_name", COMPANY_OPTIONS[0])
    st.session_state.setdefault("gcv_bank_name", BANK_OPTIONS[0])
    st.session_state.setdefault("gcv_card_type", "debit")
    st.session_state.setdefault("gcv_card_last4", "")
    st.session_state.setdefault("gcv_summary_preview_df", pd.DataFrame())


def clear_processed_results() -> None:
    st.session_state.gcv_processed = False
    st.session_state.gcv_results_saved = False
    st.session_state.gcv_summary_df = pd.DataFrame()
    st.session_state.gcv_summary_preview_df = pd.DataFrame()
    st.session_state.gcv_raw_df = pd.DataFrame()
    st.session_state.gcv_excel_bytes = b""
    st.session_state.gcv_errors = []
    st.session_state.last_database_csv_path = ""
    st.session_state.gcv_pending_uploads = []


def reset_upload_form() -> None:
    st.session_state.gcv_company_name = COMPANY_OPTIONS[0]
    st.session_state.gcv_bank_name = BANK_OPTIONS[0]
    st.session_state.gcv_card_type = "debit"
    st.session_state.gcv_card_last4 = ""


def resolve_streamlit_token() -> tuple[str | None, bool, str]:
    qp = st.query_params
    auth_code = qp.get("code")
    auth_state = qp.get("state")

    if auth_code:
        expected_state = st.session_state.get("oauth_state")
        try:
            if expected_state and auth_state != expected_state:
                raise RuntimeError("Invalid OAuth state. Please try connecting again.")
            token = finish_microsoft_redirect_flow(str(auth_code))
            st.session_state["graph_token"] = token
            st.session_state.pop("oauth_state", None)
            st.query_params.clear()
            st.success("Connected to Microsoft")
            st.rerun()
        except Exception as exc:
            st.query_params.clear()
            st.error(f"Microsoft login failed: {exc}")

    available_auth, auth_reason = _microsoft_auth_available()
    token = get_microsoft_token_silent()
    if token:
        st.session_state["graph_token"] = token
    elif "graph_token" in st.session_state:
        token = st.session_state["graph_token"]

    return token, available_auth, auth_reason


def render_auth_status(token: str | None, available_auth: bool, auth_reason: str) -> None:
    auth_col1, auth_col2, auth_col3 = st.columns([1, 2, 1])

    with auth_col1:
        if token:
            st.success("Microsoft connected")
        elif available_auth:
            st.warning("Microsoft not connected")
        else:
            st.info("Microsoft login unavailable")

    with auth_col2:
        if token:
            st.caption("Microsoft session is active for this Streamlit session.")
        elif available_auth:
            if "oauth_state" not in st.session_state:
                st.session_state["oauth_state"] = str(uuid4())
            try:
                login_url = get_microsoft_login_url(st.session_state["oauth_state"])
                st.link_button("Connect to Microsoft", login_url, type="primary")
            except Exception as exc:
                st.error(f"Could not build Microsoft login URL: {exc}")
        else:
            st.caption(auth_reason)

    with auth_col3:
        if token and st.button("Disconnect Microsoft"):
            clear_microsoft_session()
            st.query_params.clear()
            st.rerun()


def render_page_header(page_title: str) -> tuple[str | None, bool, str]:
    st.title(page_title)
    st.caption(f"Database folder: {_database_dir()}")
    initialize_session_state()
    token, available_auth, auth_reason = resolve_streamlit_token()
    render_auth_status(token, available_auth, auth_reason)
    return token, available_auth, auth_reason


def render_process_page(token: str | None) -> None:
    uploaded_pdf = st.file_uploader("Upload PDF (one invoice per page)", type=["pdf"], key="gcv_upload_pdf")
    current_upload_signature = ""
    if uploaded_pdf is not None:
        current_upload_signature = f"{uploaded_pdf.name}:{uploaded_pdf.size}"
    previous_upload_signature = str(st.session_state.get("gcv_upload_signature", ""))
    if current_upload_signature != previous_upload_signature:
        clear_processed_results()
        reset_upload_form()
        st.session_state.gcv_upload_signature = current_upload_signature
        st.rerun()

    company_name = st.selectbox("Company *", options=COMPANY_OPTIONS, index=0, key="gcv_company_name")
    bank_name = st.selectbox("Bank *", options=BANK_OPTIONS, index=0, key="gcv_bank_name")
    card_type = st.selectbox("Card type *", options=["debit", "credit"], index=0, key="gcv_card_type")
    card_last4 = st.text_input("Card last 4 digits *", max_chars=4, key="gcv_card_last4")
    process_clicked = st.button("Process", type="primary")

    if process_clicked:
        if not token:
            st.error("Microsoft connection is required to save PDFs and update the cloud CSV.")
            st.stop()
        if not uploaded_pdf:
            st.error("Please upload a PDF file.")
            st.stop()
        card_last4 = _normalize_card_last4(card_last4)
        if not card_last4:
            st.error("Card last 4 digits must be exactly 4 numbers.")
            st.stop()
        pdf_bytes = uploaded_pdf.read()
        if not pdf_bytes:
            st.error("Uploaded file appears empty.")
            st.stop()

        summary_rows: list[dict[str, Any]] = []
        raw_rows: list[dict[str, Any]] = []
        database_rows: list[dict[str, Any]] = []
        pending_uploads: list[dict[str, Any]] = []
        errors: list[str] = []
        progress = st.progress(0)
        status = st.empty()

        try:
            drive_id = resolve_drive_id(token)
            remote_names = existing_remote_file_names(token, drive_id)
            with tempfile.TemporaryDirectory(prefix="invoice_pages_gcv_") as tmp_dir:
                pages = render_pdf_pages_to_png_paths(pdf_bytes=pdf_bytes, output_dir=tmp_dir, dpi=200)
                if not pages:
                    st.error("No pages found in PDF.")
                    st.stop()

                for idx, page in enumerate(pages, start=1):
                    page_number = page["page_number"]
                    image_path = page["image_path"]
                    status.write(f"Processing page {page_number}/{len(pages)}...")

                    ocr_json: dict[str, Any] = {}
                    compact: dict[str, Any] = {}
                    gpt_json: dict[str, Any] = fallback_gpt("Page processing failed before classification")
                    error_msg = None

                    try:
                        with open(image_path, "rb") as fh:
                            image_bytes = fh.read()
                        ocr_json = extract_text_google_vision(image_bytes)
                        compact = parse_receipt_from_text(ocr_json.get("full_text", ""))
                        gpt_json = classify_with_gpt(
                            vision_json=ocr_json,
                            compact_receipt=compact,
                            use_location_enrichment=True,
                        )
                    except Exception as exc:
                        error_msg = str(exc)
                        errors.append(f"Page {page_number}: {error_msg}")
                        st.warning(f"Page {page_number} failed: {error_msg}")
                        logger.exception("Error processing page %s", page_number)

                    vision_payment_date = str(compact.get("date") or "")
                    vision_total = _to_float(compact.get("total"), 0.0)
                    vision_taxes = _to_float(compact.get("taxes_total"), 0.0)
                    vision_merchant = str(compact.get("merchant") or "").strip()
                    vision_city = str(compact.get("city") or "").strip()
                    vision_province = str(compact.get("province") or "").strip()

                    gpt_category = str(gpt_json.get("category") or "other").strip().lower()
                    if gpt_category not in ALLOWED_CATEGORIES:
                        gpt_category = "other"

                    notes = str(gpt_json.get("notes") or "").strip()
                    if error_msg:
                        notes = f"Page error: {error_msg}"

                    suggested_file_name = build_suggested_file_name(
                        payment_date=str(gpt_json.get("payment_date") or vision_payment_date),
                        bank=bank_name.strip(),
                        card_type=card_type,
                        merchant_name=str(gpt_json.get("merchant_name") or vision_merchant),
                        total_amount=_to_float(gpt_json.get("total_amount"), vision_total),
                    )
                    final_pdf_name, final_pdf_path = make_unique_remote_pdf_name(
                        suggested_file_name or f"page_{page_number}",
                        remote_names,
                    )
                    final_pdf_bytes = extract_pdf_page_bytes(pdf_bytes=pdf_bytes, page_index=page_number - 1)
                    pending_uploads.append(
                        {
                            "file_name": final_pdf_name,
                            "file_path": final_pdf_path,
                            "content": final_pdf_bytes,
                        }
                    )

                    summary_rows.append(
                        {
                            "gpt_payment_date": str(gpt_json.get("payment_date") or ""),
                            "gpt_total_amount": _to_float(gpt_json.get("total_amount"), 0.0),
                            "gpt_taxes_total": _to_float(gpt_json.get("taxes_total"), 0.0),
                            "gpt_category": gpt_category,
                            "gpt_merchant_name": str(gpt_json.get("merchant_name") or ""),
                            "gpt_city": str(gpt_json.get("city") or ""),
                            "gpt_province": str(gpt_json.get("province") or ""),
                            "gpt_confidence": _to_float(gpt_json.get("confidence"), 0.0),
                            "company": company_name,
                            "bank": bank_name.strip(),
                            "card_type": card_type,
                            "card_last4": card_last4,
                            "file_name": final_pdf_name,
                            "file_path": final_pdf_path,
                            "source_page_number": page_number,
                            "notes": notes,
                        }
                    )

                    database_rows.append(
                        build_database_row(
                            page_number=page_number,
                            company=company_name,
                            bank=bank_name.strip(),
                            card_type=card_type,
                            card_last4=card_last4,
                            gpt_json=gpt_json,
                            notes=notes,
                            pdf_file_name=final_pdf_name,
                            pdf_file_path=final_pdf_path,
                        )
                    )

                    raw_rows.append(
                        {
                            "source_page_number": page_number,
                            "raw_google_vision_json": _safe_json(ocr_json),
                            "raw_gpt_json": _safe_json(gpt_json),
                            "vision_payment_date": vision_payment_date,
                            "vision_total_amount": vision_total,
                            "vision_taxes_total": vision_taxes,
                            "vision_merchant_name": vision_merchant,
                            "vision_city": vision_city,
                            "vision_province": vision_province,
                        }
                    )
                    progress.progress(idx / len(pages))
        except Exception as exc:
            st.error(f"Fatal error: {exc}")
            logger.exception("Fatal processing error")
            st.stop()

        summary_df = pd.DataFrame(summary_rows)
        raw_df = pd.DataFrame(raw_rows)
        database_df = pd.DataFrame(database_rows)
        ordered_cols = [
            "processed_at",
            "company",
            "bank",
            "card_type",
            "card_last4",
            "gpt_payment_date",
            "gpt_total_amount",
            "gpt_taxes_total",
            "gpt_category",
            "gpt_merchant_name",
            "gpt_city",
            "gpt_province",
            "gpt_confidence",
            "notes",
            "file_name",
            "file_path",
            "source_page_number",
        ]
        summary_df = database_df.reindex(columns=ordered_cols)

        st.session_state.gcv_summary_df = summary_df
        preview_df = _prettify_dataframe_columns(summary_df)
        st.session_state.gcv_summary_preview_df = preview_df
        st.session_state.gcv_raw_df = raw_df
        st.session_state.gcv_excel_bytes = create_excel_bytes(summary_df, raw_df)
        st.session_state.gcv_errors = errors
        st.session_state.gcv_pending_uploads = pending_uploads
        st.session_state.gcv_processed = True
        st.session_state.gcv_results_saved = False
        st.session_state.last_database_csv_path = ""

    if st.session_state.gcv_processed:
        st.subheader("Summary Preview")
        st.dataframe(
            st.session_state.gcv_summary_preview_df,
            use_container_width=True,
            hide_index=True,
            column_config={"Card Last4": st.column_config.TextColumn("Card Last4")},
        )
        if st.session_state.gcv_results_saved and st.session_state.last_database_csv_path:
            st.caption(f"CSV updated: {st.session_state.last_database_csv_path}")
        elif not st.session_state.gcv_results_saved:
            st.info("Results are pending review. Use Keep Results to save them or Drop Results to discard them.")
        with st.expander("Raw OCR and GPT outputs"):
            st.dataframe(st.session_state.gcv_raw_df, use_container_width=True)
        st.download_button(
            label="Download Excel Summary",
            data=st.session_state.gcv_excel_bytes,
            file_name="invoice_summary_google_vision.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        if not st.session_state.gcv_results_saved:
            action_col1, action_col2 = st.columns(2)
            with action_col1:
                if st.button("Keep Results", type="primary", use_container_width=True):
                    try:
                        drive_id = resolve_drive_id(token)
                        for pending_upload in st.session_state.gcv_pending_uploads:
                            upload_sharepoint_file_bytes(
                                str(pending_upload["file_path"]),
                                pending_upload["content"],
                                token,
                                drive_id=drive_id,
                            )
                        csv_path = append_rows_to_database(st.session_state.gcv_summary_df, token, drive_id)
                        st.session_state.last_database_csv_path = csv_path
                        st.session_state.gcv_results_saved = True
                        st.session_state.gcv_pending_uploads = []
                        st.rerun()
                    except Exception as exc:
                        st.error(f"Could not keep results: {exc}")
                        logger.exception("Could not persist processed results")
            with action_col2:
                if st.button("Drop Results", use_container_width=True):
                    clear_processed_results()
                    st.rerun()
        if st.session_state.gcv_errors:
            st.error(f"Completed with {len(st.session_state.gcv_errors)} page error(s).")
        else:
            st.success("All pages processed successfully.")


def render_database_page(token: str | None) -> None:
    if not token:
        st.info("Connect to Microsoft to browse the cloud receipts database.")
        return

    drive_id = resolve_drive_id(token)
    database_df = load_database_df(token, drive_id)
    if database_df.empty:
        st.info("No rows found in the receipts database CSV yet.")
        return

    render_database_browser(database_df, token, drive_id, section_key="database")
